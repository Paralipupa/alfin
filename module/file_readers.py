import abc
import csv
import os
import pathlib
import logging
import xlrd
import xlwt
import math
import pandas as pd
import shutil
from datetime import datetime
from zipfile import ZipFile
from openpyxl import load_workbook
from module.settings import BASE_DIR


logger = logging.getLogger(__name__)


def fatal_error(func):
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except Exception as ex:
            logger.exception("Fatal error")
            exit()

    return wrapper


def warning_error(func):
    def wrapper(*args):
        try:
            return func(*args)
        except Exception as ex:
            logger.exception("Warning error")
            return None

    return wrapper


def rchop(s, sub):
    return s[: -len(sub)] if s.endswith(sub) else s


@warning_error
def import_1c(filename: str) -> str:
    name_tmp = f'tmp_{datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}'
    tmp_folder = "tmp"
    os.makedirs(tmp_folder, exist_ok=True)

    # Распаковываем excel как zip в нашу временную папку
    with ZipFile(filename) as excel_container:
        excel_container.extractall(tmp_folder)

    # Переименовываем файл с неверным названием
    wrong_file_path = os.path.join(tmp_folder, "xl", "SharedStrings.xml")
    if not os.path.isfile(wrong_file_path):
        shutil.rmtree(tmp_folder)
        return filename
    correct_file_path = os.path.join(tmp_folder, "xl", "sharedStrings.xml")
    os.rename(wrong_file_path, correct_file_path)

    # Запаковываем excel обратно в zip и переименовываем в исходный файл
    # shutil.make_archive(pathlib.Path(os.path.dirname(filename),f'{name_tmp}'), 'zip', os.path.dirname(filename))
    shutil.make_archive(
        pathlib.Path(os.path.dirname(filename), name_tmp), "zip", tmp_folder
    )
    shutil.rmtree(tmp_folder)
    os.remove(filename)
    os.rename(pathlib.Path(os.path.dirname(filename), f"{name_tmp}.zip"), filename)
    return filename


class DataFile(abc.ABC):
    def __init__(self, fname, sheet_name, first_line: int = 0, columns: range = [50]):
        self._fname = fname
        self._first_line = first_line
        self._sheet_name = sheet_name
        self._columns = columns

    def __iter__(self):
        return self

    def __next__(self):
        return ""


# %%


class CsvFile(DataFile):
    def __init__(self, fname, first_line, columns, page_index=None):
        super(CsvFile, self).__init__(fname, "", first_line, range(columns))
        self._freader = open(fname, "r", encoding="windows-1251")
        self._first_line = first_line
        self._reader = csv.reader(self._freader, delimiter=";", quotechar="|")
        self._line_num = 0

    def get_row(self, row):
        index = 0
        for cell in row:
            if index in self._columns:
                yield XlsFile.get_cell_text(cell)
            index = index + 1

    def __next__(self):
        for row in self._reader:
            self._line_num += 1
            if self._line_num < self._first_line:
                continue
            return row
        raise StopIteration

    def __del__(self):
        self._freader.close()


# %%


class PandasFile(DataFile):
    def __init__(
        self, fname, sheet_name, first_line: int = 0, columns: int = 50, page_index=0
    ):
        super(PandasFile, self).__init__(fname, sheet_name, first_line, range(columns))
        if self._sheet_name:
            self._sheet = next(
                iter(pd.read_excel(fname, sheet_name=[self._sheet_name]).values())
            )
            xlsx = pd.ExcelFile(fname)
            self._sheet = xlsx.parse(sheet_name=[self._sheet_name])
        else:
            import_1c(fname)
            # xlsx = pd.ExcelFile(fname)
            # self._sheet = xlsx.parse(sheet_name=None)
            self._sheet = next(iter(pd.read_excel(fname, sheet_name=None).values()))
        self._rows = (row for index, row in self._sheet.iterrows())
        self._cc = 0

    @staticmethod
    def get_cell_text(cell):
        if isinstance(cell, float):
            return str(cell) if not math.isnan(cell) else ""
        return str(cell)

    def get_row(self, row):
        index = 0
        for cell in row:
            if index in self._columns:
                yield PandasFile.get_cell_text(cell)
            index = index + 1

    def __next__(self):
        for row in self._rows:
            return list(self.get_row(row))
        raise StopIteration

    def __del__(self):
        pass


# %%


class XlsFile(DataFile):
    def __init__(
        self, fname, sheet_name, first_line: int = 0, columns: int = 50, page_index=0
    ):
        super(XlsFile, self).__init__(fname, sheet_name, first_line, range(columns))
        self._book = xlrd.open_workbook(
            fname, logfile=open(os.devnull, "w"), ignore_workbook_corruption=True
        )
        if self._sheet_name:
            sheet = self._book.sheet_by_name(self._sheet_name)
        else:
            sheet = self._book.sheets()[page_index]
        self._rows = (sheet.row(index) for index in range(first_line, sheet.nrows))

    @staticmethod
    def get_cell_text(cell):
        if cell.ctype == 2:
            return rchop(str(cell.value), ".0")
        return str(cell.value)

    def get_row(self, row):
        index = 0
        for cell in row:
            if index in self._columns:
                yield XlsFile.get_cell_text(cell)
            index = index + 1

    def __next__(self):
        for row in self._rows:
            return list(self.get_row(row))
        raise StopIteration

    def __del__(self):
        pass


# %%


class XlsxFile(DataFile):
    @fatal_error
    def __init__(
        self,
        fname,
        sheet_name,
        first_line: int = 0,
        columns: int = 50,
        page_index: int = 0,
    ):
        super(XlsxFile, self).__init__(fname, sheet_name, first_line, range(columns))

        file_name = import_1c(fname)
        self._wb = load_workbook(filename=file_name, read_only=True, data_only=True)
        if "TDSheet" in self._wb.sheetnames:
            self._sheet_name = "TDSheet"
        if self._sheet_name:
            self._ws = self._wb.get_sheet_by_name(self._sheet_name)
        else:
            self._ws = self._wb.worksheets[page_index]
        self._cursor = self._ws.iter_rows()
        row_num = 0
        while row_num < self._first_line:
            row_num += 1
            next(self._cursor)

    @staticmethod
    def get_cell_text(cell):
        return str(cell.value) if cell.value else ""

    def get_row(self, row):
        i = 0
        for cell in row:
            if i in self._columns:
                yield XlsxFile.get_cell_text(cell)
            i += 1

    def __next__(self):
        return list(self.get_row(next(self._cursor)))

    def __del__(self):
        self._wb.close()

    def get_index(self, cell):
        try:
            return cell.column
        except AttributeError:
            return -1


# %%


class XlsWrite:
    def __init__(self, filename: str):
        self.name = filename
        self.book = xlwt.Workbook(encoding="utf-8", style_compression=2)

    def save(self) -> str:
        try:
            path = os.path.join(BASE_DIR, "output")
            pathlib.Path.mkdir(pathlib.Path(path), exist_ok=True)
            i = 0
            name = self.name
            while os.path.isfile(pathlib.Path(path, f"{name}.xls")) and i < 100:
                i += 1
                name = f"{self.name}({i})"
            self.book.save(pathlib.Path(path, f"{name}.xls"))
            return pathlib.Path(path, f"{name}.xls")
        except Exception as ex:
            logger.exception("Save error")
        return None

    def addSheet(self, title: str = ""):
        title = f"Лист {len(self.book._Workbook__worksheets)+1}" if not title else title
        self.sheet = self.book.add_sheet(title)
        return self.sheet

    def write(
        self,
        row: int,
        col: int,
        value,
        style_string: str = None,
        type_name: str = None,
        num_format_str: str = None,
    ):
        try:
            if num_format_str is None and type_name is not None and type_name == "date":
                num_format_str = r"dd/mm/yyyy"
            if isinstance(value, str):
                neededWidth = int((1 + min([len(str(value)), 64])) * 256)
            else:
                neededWidth = 12 * 256
            if style_string or num_format_str:
                if style_string and not num_format_str:
                    style = xlwt.easyxf(style_string)
                elif not style_string and num_format_str:
                    style = xlwt.easyxf(num_format_str=num_format_str)
                else:
                    style = xlwt.easyxf(style_string, num_format_str=num_format_str)
                self.sheet.write(row, col, value, style=style)
            else:
                self.sheet.write(row, col, value)
            if self.sheet.col(col).width < neededWidth:
                self.sheet.col(col).width = neededWidth
        except ValueError as ex:
            logger.exception(f"Write to Excel:row={row} col={col} value={value}")
        except Exception as ex:
            logger.exception(f"Write to Excel:row={row} col={col} value={value}")


def get_file_reader(fname):
    """Get class for reading file as iterable"""
    _, file_extension = os.path.splitext(fname)
    if file_extension == ".xls":
        # return PandasFile
        return XlsFile
    if file_extension == ".xlsx":
        # return PandasFile
        return XlsxFile
    if file_extension == ".csv":
        return CsvFile
    raise Exception(f"Unknown file type {file_extension}")


def get_file_write(fname):
    return XlsWrite
